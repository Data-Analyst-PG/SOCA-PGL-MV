# portal_app/modules/gestion_solicitudes/complementarias.py
# ─────────────────────────────────────────────────────────────────────────────
# Vista de GESTIÓN de complementarias (auditor / admin).
# Sin HTML propio — visual viene de ui/components.
# Modal nativo (@st.dialog) para edición.
#
# Estructura:
#   ① KPIs globales
#   ② Solicitudes ACTIVAS (Pendiente + En revisión) — cards editables
#   ③ Historial completo — tabla con filtros + descarga Excel
# ─────────────────────────────────────────────────────────────────────────────
from __future__ import annotations
from collections import Counter
from io import BytesIO

import pandas as pd
import streamlit as st

from ui.components import (
    section_header, 
    kpi_row, alert,
    solicitud_card, 
    historial_timeline, 
    status_badge_html,
    solicitudes_table, 
    page_banner,
)
from .shared import get_complementarias, update_complementaria, now_iso_utc, log_accion_complementarias as log_accion
from services.notificaciones import enviar_notificacion
from services.access import check_access
from services.supabase_client import get_authed_client


def _obtener_reglas_alcance(user_id: str) -> list[dict]:
    """Reglas de alcance activas para este usuario. Lista vacía = sin alcance
    asignado (no debería ver nada, salvo que tenga acceso total)."""
    try:
        supabase = get_authed_client()
        res = (
            supabase.table("alcance_usuario")
            .select("empresa,tipo")
            .eq("user_id", user_id)
            .eq("modulo", "complementarias")
            .eq("activo", True)
            .execute()
        )
        # Homologar nombre de columna: tipo → tipo_complementaria
        return [
            {"empresa": r["empresa"], "tipo_complementaria": r["tipo"]}
            for r in (res.data or [])
        ]
    except Exception:
        return []


def _filtrar_por_alcance(comps: list[dict], user_id: str) -> list[dict]:
    """Filtra la lista de complementarias según el alcance del usuario.
    Acceso total (complementarias:ver_todo) → ve todo, sin filtrar."""
    if check_access(user_id, None, "complementarias:ver_todo"):
        return comps

    reglas = _obtener_reglas_alcance(user_id)
    if not reglas:
        return []  # sin reglas asignadas = no ve nada (default seguro)

    permitidos = []
    for c in comps:
        emp  = c.get("empresa")
        tipo = c.get("tipo_complementaria")
        for r in reglas:
            r_emp  = r.get("empresa")
            r_tipo = r.get("tipo_complementaria")
            if (r_emp is None or r_emp == emp) and (r_tipo is None or r_tipo == tipo):
                permitidos.append(c)
                break
    return permitidos


# ── Catálogos ─────────────────────────────────────────────────────────────────
ESTATUSES        = ["Pendiente", "En revisión", "Resuelto", "Cancelado"]
ESTATUSES_ACTIVOS = {"Pendiente", "En revisión"}

# Columnas para la tabla / Excel
COLS_TABLA = [
    ("folio",                  "Folio"),
    ("fecha_captura",          "Fecha captura"),
    ("empresa",                "Empresa"),
    ("sucursal",               "Sucursal"),
    ("plataforma",             "Plataforma"),
    ("solicitante",            "Solicitante"),
    ("correo",                 "Correo"),
    ("numero_trafico",         "Tráfico"),
    ("tipo_complementaria",    "Tipo"),
    ("tipo_motivo",            "Tipo de motivo"),
    ("estatus",                "Estatus"),
    ("auditor",                "Auditor"),
    ("fecha_resuelto",         "Fecha resolución"),
    ("comentarios_auditor",    "Comentarios auditor"),
]

# ── Adjuntar captura de conrimación ───────────────────────────────────────────
def _preparar_adjunto_captura(archivo, folio_fmt: str) -> dict | None:
    """Prepara la captura de 'Resuelto' para adjuntarse al correo — NO se
    guarda en Supabase Storage. Comprime la imagen si excede 5 MB."""
    try:
        extension = archivo.name.rsplit(".", 1)[-1].lower()

        if extension == "pdf":
            file_bytes = archivo.read()
            nombre = f"confirmacion_{folio_fmt}.pdf"
            return {"filename": nombre, "content_bytes": file_bytes}

        from PIL import Image
        from io import BytesIO

        img = Image.open(archivo).convert("RGB")

        # Primer intento: ancho máx 1200px, calidad 70
        max_width = 1200
        if img.width > max_width:
            ratio = max_width / img.width
            img = img.resize((max_width, int(img.height * ratio)), Image.LANCZOS)
        buffer = BytesIO()
        img.save(buffer, format="JPEG", quality=70, optimize=True)
        file_bytes = buffer.getvalue()

        # Si sigue pesando más de 5 MB, segunda pasada más agresiva
        if len(file_bytes) > 5 * 1024 * 1024:
            if img.width > 800:
                ratio = 800 / img.width
                img = img.resize((800, int(img.height * ratio)), Image.LANCZOS)
            buffer = BytesIO()
            img.save(buffer, format="JPEG", quality=50, optimize=True)
            file_bytes = buffer.getvalue()

        nombre = f"confirmacion_{folio_fmt}.jpg"
        return {"filename": nombre, "content_bytes": file_bytes}
    except Exception:
        return None
        
# ── Excel ─────────────────────────────────────────────────────────────────────
def _to_excel(df: pd.DataFrame) -> bytes:
    buf = BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Complementarias")
        ws = writer.sheets["Complementarias"]
        for col_cells in ws.columns:
            max_len = max(
                len(str(c.value)) if c.value is not None else 0
                for c in col_cells
            )
            ws.column_dimensions[col_cells[0].column_letter].width = min(max_len + 4, 50)
        from openpyxl.styles import Font, PatternFill, Alignment
        fill = PatternFill("solid", fgColor="1B2266")
        font = Font(color="FFFFFF", bold=True)
        for cell in ws[1]:
            cell.fill      = fill
            cell.font      = font
            cell.alignment = Alignment(horizontal="center", vertical="center")
    return buf.getvalue()



def _get_factura_url(factura_path: str) -> str | None:
    """Genera la URL pública del archivo en Supabase Storage."""
    try:
        from services.supabase_client import get_authed_client
        supabase = get_authed_client()
        res = supabase.storage.from_("complementarias-evidencias").get_public_url(factura_path)
        return res
    except Exception:
        return None


# ── Modal de edición ──────────────────────────────────────────────────────────
@st.dialog("Actualizar complementaria", width="large")
def _modal_edicion(comp: dict, gestor: str, solo_lectura: bool = False):
    folio = comp.get("folio")
    est   = comp.get("estatus", "Pendiente")
    badge = status_badge_html(est)
    tipo  = comp.get("tipo_complementaria", "")
    es_desconclusion = tipo == "Desconclusión"

    st.markdown(
        f"**Folio {int(folio):04d}** — {tipo}&nbsp;&nbsp;{badge}",
        unsafe_allow_html=True,
    )
    col1, col2, col3 = st.columns(3)
    col1.markdown(f"**👤 Solicitante:** {comp.get('solicitante','')}")
    col2.markdown(f"**🚛 Tráfico:** {comp.get('numero_trafico','—')}")
    col3.markdown(f"**📅 Capturada:** {str(comp.get('fecha_captura',''))[:10]}")
    st.caption(
        f"🏢 {comp.get('empresa','')}  |  "
        f"📍 {comp.get('sucursal','')}  |  "
        f"💻 {comp.get('plataforma','')}"
    )

    st.divider()

    # ── Motivo ────────────────────────────────────────────────────────────────
    motivo = comp.get("motivo_solicitud", "")
    tipo_motivo = comp.get("tipo_motivo", "")
    if tipo_motivo:
        st.markdown(f"**Tipo de motivo:** {tipo_motivo}")
    if motivo:
        st.markdown(f"**Motivo de la solicitud:** {motivo}")

    # ── Detalle del concepto ──────────────────────────────────────────────────
    if not es_desconclusion:
        st.divider()
        st.markdown("**📦 Detalle de la solicitud:**")

        def _fila(label, valor):
            if valor not in [None, "", "N/A", "Sin datos"]:
                st.markdown(f"- **{label}:** {valor}")

        if tipo == "Modificación":
            ca, cn = st.columns(2)
            with ca:
                st.markdown("**📝 Datos actuales**")
                _fila("Tipo concepto",  comp.get("tipo_concepto_actual"))
                _fila("Concepto",       comp.get("concepto_actual"))
                _fila("Proveedor",      comp.get("proveedor_actual"))
                _fila("Moneda",         comp.get("moneda_actual"))
                imp_a = comp.get("importe_actual")
                if imp_a is not None:
                    st.markdown(f"- **Importe:** ${float(imp_a):,.2f}")
                _fila("Tasa IVA",       comp.get("tasa_iva_actual"))
                _fila("Tasa Retención", comp.get("tasa_retencion_actual"))
                _fila("Retención ISR",  comp.get("retencion_isr_actual"))
                tot_a = comp.get("total_actual")
                if tot_a is not None:
                    st.markdown(f"- **Total calculado:** ${float(tot_a):,.2f}")
            with cn:
                st.markdown("**✅ Datos correctos**")
                _fila("Tipo concepto",  comp.get("tipo_concepto_nuevo"))
                _fila("Concepto",       comp.get("concepto_nuevo"))
                _fila("Proveedor",      comp.get("proveedor_nuevo"))
                _fila("Moneda",         comp.get("moneda_nuevo"))
                imp_n = comp.get("importe_nuevo")
                if imp_n is not None:
                    st.markdown(f"- **Importe:** ${float(imp_n):,.2f}")
                _fila("Tasa IVA",       comp.get("tasa_iva_nuevo"))
                _fila("Tasa Retención", comp.get("tasa_retencion_nuevo"))
                _fila("Retención ISR",  comp.get("retencion_isr_nuevo"))
                tot_n = comp.get("total_nuevo")
                if tot_n is not None:
                    st.markdown(f"- **Total calculado:** ${float(tot_n):,.2f}")
        else:
            # Agregar Concepto — solo lado nuevo
            st.markdown("**✅ Concepto a agregar**")
            _fila("Tipo concepto",  comp.get("tipo_concepto_nuevo"))
            _fila("Concepto",       comp.get("concepto_nuevo"))
            _fila("Proveedor",      comp.get("proveedor_nuevo"))
            _fila("Moneda",         comp.get("moneda_nuevo"))
            imp_n = comp.get("importe_nuevo")
            if imp_n is not None:
                st.markdown(f"- **Importe:** ${float(imp_n):,.2f}")
            _fila("Tasa IVA",       comp.get("tasa_iva_nuevo"))
            _fila("Tasa Retención", comp.get("tasa_retencion_nuevo"))
            _fila("Retención ISR",  comp.get("retencion_isr_nuevo"))
            tot_n = comp.get("total_nuevo")
            if tot_n is not None:
                st.markdown(f"- **Total calculado:** ${float(tot_n):,.2f}")

    # ── Factura adjunta ───────────────────────────────────────────────────────
    factura_path = comp.get("factura_path")
    if factura_path:
        st.divider()
        st.markdown("**📎 Factura adjunta:**")
        url = _get_factura_url(factura_path)
        if url:
            st.link_button("📄 Ver factura", url, use_container_width=False)
        else:
            st.caption("No se pudo generar el enlace de la factura.")
    
    st.divider()
    st.markdown("**📋 Historial:**")
    historial_timeline(comp.get("historial") or [])
    st.divider()

    if solo_lectura:
        col_x, _ = st.columns([1, 3])
        with col_x:
            if st.button("Cerrar", key=f"gc_comp_close_{folio}"):
                st.rerun()
        return

    st.markdown("**✏️ Registrar actualización:**")

    col_a, col_b = st.columns(2)
    with col_a:
        idx_est    = ESTATUSES.index(est) if est in ESTATUSES else 0
        nuevo_est  = st.selectbox("Nuevo estatus", ESTATUSES, index=idx_est, key=f"gc_comp_est_{folio}")
    with col_b:
        aud_actual = comp.get("auditor") or "Sin asignar"
        nuevo_aud  = gestor
        st.text_input("Auditor", value=gestor, disabled=True,
                      help="Se asigna automáticamente al usuario que gestiona la solicitud.")
    comentario = st.text_area(
        "Comentario del auditor",
        value=comp.get("comentarios_auditor") or "",
        placeholder="Observaciones, resolución, motivo de cancelación...",
        height=100,
        key=f"gc_comp_com_{folio}",
    )

    captura_resuelto = None
    if nuevo_est == "Resuelto" and tipo != "Desconclusión":
        captura_resuelto = st.file_uploader(
            "📸 Captura de confirmación (opcional)",
            type=["png", "jpg", "jpeg", "pdf"],
            key=f"gc_comp_captura_{folio}",
            help="Se adjunta al correo de notificación — no se guarda en el sistema.",
        )

    col_b1, col_b2 = st.columns([1, 3])
    with col_b1:
        guardar = st.button("💾 Guardar", type="primary", key=f"gc_comp_save_{folio}")
    with col_b2:
        if st.button("✖️ Cancelar", key=f"gc_comp_cancel_{folio}"):
            st.rerun()

    if guardar:
        cambios_texto = []
        if nuevo_est != est:
            cambios_texto.append(f"Estatus: {est} → {nuevo_est}")
        if nuevo_aud != aud_actual:
            cambios_texto.append(f"Auditor: {aud_actual} → {nuevo_aud}")
        if comentario.strip() != (comp.get("comentarios_auditor") or "").strip():
            cambios_texto.append(f"Comentario actualizado")

        if not cambios_texto:
            st.warning("No hay cambios que guardar.")
            return

        now = now_iso_utc()
        historial = list(comp.get("historial") or [])
        historial.append({
            "at":      now,
            "by":      gestor,
            "action":  "update",
            "details": " | ".join(cambios_texto),
        })

        changes = {
            "fecha_ultima_modificacion": now,
            "auditor":               nuevo_aud,
            "comentarios_auditor":   comentario.strip(),
            "historial":             historial,
        }
        if nuevo_est != est:
            changes["estatus"] = nuevo_est
            if nuevo_est == "Resuelto":
                changes["fecha_resuelto"] = now[:10]

        if update_complementaria(folio, changes):
            log_accion("editar_solicitud", {"folio": folio, "estatus": nuevo_est})

            # ── Notificación: en CUALQUIER cambio de estatus ────────────────────
            hubo_cambio_estatus = nuevo_est != est
            correo_enviado = None
            if hubo_cambio_estatus:
                folio_fmt = f"{int(folio):04d}"
                _colores_estatus = {
                    "Pendiente":    ("#DBEAFE", "#1E40AF"),
                    "En revisión":  ("#FEF3C7", "#92400E"),
                    "Resuelto":     ("#D1FAE5", "#065F46"),
                    "Cancelado":    ("#FEE2E2", "#7F1D1D"),
                }
                color_bg, color_fg = _colores_estatus.get(nuevo_est, ("#F3F4F6", "#374151"))
                comentario_correo = comentario.strip() or "Sin comentarios adicionales."

                adjunto_captura = None
                if nuevo_est == "Resuelto" and captura_resuelto is not None:
                    adjunto_captura = _preparar_adjunto_captura(captura_resuelto, folio_fmt)

                resultado_correo = enviar_notificacion(
                    modulo="complementarias",
                    evento="estatus_actualizado",
                    folio=folio_fmt,
                    clave_unica=f"{nuevo_est}_{now}",
                    datos={
                        "solicitante": comp.get("solicitante", ""),
                        "empresa": comp.get("empresa", ""),
                        "numero_trafico": comp.get("numero_trafico", ""),
                        "tipo": comp.get("tipo_complementaria", ""),
                        "auditor": nuevo_aud,
                        "estatus": nuevo_est,
                        "comentario": comentario_correo,
                        "color_bg": color_bg,
                        "color_fg": color_fg,
                    },
                    tipo_solicitud=comp.get("tipo_complementaria"),
                    empresa=comp.get("empresa"),
                    correo_solicitante=comp.get("correo"),
                    adjunto=adjunto_captura,
                )
                correo_enviado = resultado_correo.get("ok", False)

            st.session_state["gc_comp_success_payload"] = {
                "folio": folio,
                "hubo_cambio_estatus": hubo_cambio_estatus,
                "correo_enviado": correo_enviado,
            }
            st.cache_data.clear()
            st.rerun()


# ── Render principal ──────────────────────────────────────────────────────────
def render():
    page_banner("📋", "Gestión de Complementarias",
                   "Revisa y actualiza solicitudes de complementarias")

    perfil = st.session_state.get("user_profile") or {}
    from services.supabase_client import current_user
    u       = current_user() or {}
    gestor  = perfil.get("full_name") or u.get("email", "Auditor")
    user_id = u.get("id") or u.get("sub") or ""

    comps = get_complementarias()
    comps = _filtrar_por_alcance(comps, user_id)

    if not comps:
        alert("info", "No hay solicitudes que te correspondan revisar por el momento.")
        return

    # ── ① KPIs ───────────────────────────────────────────────────────────────
    conteo = Counter(c.get("estatus", "Pendiente") for c in comps)
    kpi_row([
        dict(icono="🕐", label="Pendientes",
             valor=conteo.get("Pendiente", 0), color="#1D4ED8"),
        dict(icono="🔍", label="En revisión",
             valor=conteo.get("En revisión", 0), color="#D97706"),
        dict(icono="✅", label="Resueltas",
             valor=conteo.get("Resuelto", 0), color="#059669"),
        dict(icono="🚫", label="Canceladas",
             valor=conteo.get("Cancelado", 0), color="#DC2626"),
    ])

    st.divider()

    # ── Diálogo de confirmación tras guardar ────────────────────────────────
    if st.session_state.get("gc_comp_success_payload"):
        payload = st.session_state["gc_comp_success_payload"]

        @st.dialog("✅ Cambios guardados")
        def _dlg_guardado():
            st.success(f"Los cambios del folio **#{int(payload['folio']):04d}** se guardaron correctamente.")
            if payload["hubo_cambio_estatus"]:
                if payload["correo_enviado"]:
                    st.info("📧 Se envió la notificación por correo automáticamente.")
                else:
                    st.warning(
                        "⚠️ Los cambios se guardaron, pero la notificación por correo "
                        "no pudo enviarse."
                    )
            if st.button("OK", type="primary", key="gc_comp_success_ok"):
                st.session_state.pop("gc_comp_success_payload", None)
                st.rerun()

        _dlg_guardado()

    # ── Abrir modal si hay selección ──────────────────────────────────────────
    if "gc_comp_modal_folio" in st.session_state:
        folio_sel = st.session_state.pop("gc_comp_modal_folio")
        comp_sel  = next((c for c in comps if c.get("folio") == folio_sel), None)
        if comp_sel:
            _modal_edicion(comp_sel, gestor)

    # ── ② ACTIVAS ─────────────────────────────────────────────────────────────
    activas = [c for c in comps if c.get("estatus", "Pendiente") in ESTATUSES_ACTIVOS]

    if activas:
        st.markdown(f"#### 🔄 Solicitudes activas ({len(activas)})")
        for c in activas:
            folio = c.get("folio", "")
            est   = c.get("estatus", "Pendiente")
            fecha = str(c.get("fecha_captura", ""))[:10]

            meta = [
                ("🏢", c.get("empresa", "")),
                ("🔖", c.get("tipo_complementaria", "")),
                ("📋", f"Tráfico: {c.get('numero_trafico','—')}"),
                ("👤 Sol.", c.get("solicitante", "")),
                ("🔍", c.get("auditor") or "Sin asignar"),
            ]

            clicked = solicitud_card(
                id_label=f"Folio {int(folio):04d}" if folio else "—",
                titulo=c.get("tipo_complementaria") or "(Sin tipo)",
                fecha=fecha,
                estatus=est,
                meta=meta,
                on_edit_key=f"gc_comp_open_{folio}",
            )
            if clicked:
                st.session_state["gc_comp_modal_folio"] = folio
                st.rerun()
    else:
        st.info("No hay solicitudes activas en este momento. ✅")

    st.divider()

    # ── ③ HISTORIAL COMPLETO ──────────────────────────────────────────────────
    st.markdown("#### 📋 Historial completo")
    st.caption("Incluye solicitudes cerradas (Resuelto / Cancelado). Ajusta los filtros y genera el reporte.")

    col1, col2, col3, col4 = st.columns(4)
    with col1:
        filtro_est = st.selectbox(
            "Estatus", ["Todos"] + ESTATUSES, key="gc_comp_h_est"
        )
    with col2:
        todas_emp = ["Todas"] + sorted(set(
            c.get("empresa", "") for c in comps if c.get("empresa")
        ))
        filtro_emp = st.selectbox("Empresa", todas_emp, key="gc_comp_h_emp")
    with col3:
        filtro_desde = st.date_input("Desde", value=None, key="gc_comp_h_desde")
    with col4:
        filtro_q = st.text_input(
            "Tráfico o solicitante", placeholder="Buscar...", key="gc_comp_h_q"
        )

    col_btn, _ = st.columns([1, 3])
    with col_btn:
        generar = st.button("🔎 Generar reporte", type="primary",
                             key="gc_comp_h_generar", use_container_width=True)

    if generar:
        filtrados = list(comps)
        if filtro_est != "Todos":
            filtrados = [c for c in filtrados if c.get("estatus") == filtro_est]
        if filtro_emp != "Todas":
            filtrados = [c for c in filtrados if c.get("empresa", "") == filtro_emp]
        if filtro_desde:
            filtrados = [
                c for c in filtrados
                if str(c.get("fecha_captura", ""))[:10] >= str(filtro_desde)
            ]
        if filtro_q.strip():
            q = filtro_q.strip().lower()
            filtrados = [
                c for c in filtrados
                if q in str(c.get("numero_trafico", "")).lower()
                or q in str(c.get("solicitante", "")).lower()
            ]
        st.session_state["gc_comp_h_resultados"] = filtrados

    resultados = st.session_state.get("gc_comp_h_resultados")

    if resultados is None:
        st.info("Aplica tus filtros y presiona **Generar reporte** para consultar el historial.")
    elif not resultados:
        st.info("No hay resultados con los filtros aplicados.")
    else:
        st.caption(f"Mostrando {len(resultados)} de {len(comps)} solicitud(es) · haz clic en una fila para ver el detalle")

        df = pd.DataFrame([
            {label: c.get(col, "") for col, label in COLS_TABLA}
            for c in resultados
        ])
        for fecha_col in ["Fecha captura", "Fecha resolución"]:
            if fecha_col in df.columns:
                df[fecha_col] = df[fecha_col].astype(str).str[:10]

        evento = st.dataframe(
            df,
            use_container_width=True,
            hide_index=True,
            on_select="rerun",
            selection_mode="single-row",
            key="gc_comp_h_tabla",
        )

        filas_sel = evento.selection.rows if evento and evento.selection else []
        if filas_sel:
            comp_sel = resultados[filas_sel[0]]
            _modal_edicion(comp_sel, gestor, solo_lectura=True)

        excel_bytes = _to_excel(df)
        descargado_excel = st.download_button(
            "⬇️ Descargar Excel",
            data=excel_bytes,
            file_name="complementarias.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="gc_comp_dl_excel",
        )
        if descargado_excel:
            log_accion("exportar_excel", {"filas": len(df)})
