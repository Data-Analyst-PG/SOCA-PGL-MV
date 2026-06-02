# portal_app/modules/gestion_solicitudes/tickets.py
# ─────────────────────────────────────────────────────────────────────────────
# Vista de GESTIÓN de tickets (admin / equipo de análisis).
#
# Estructura:
#   ① KPIs globales
#   ② Tickets ACTIVOS (sin Concluido/Cancelado) — cards editables
#   ③ Historial completo — tabla con filtros múltiples + descarga Excel
# ─────────────────────────────────────────────────────────────────────────────
from __future__ import annotations
from collections import Counter
from datetime import datetime, timezone
from io import BytesIO

import pandas as pd
import streamlit as st

from services.supabase_client import current_user, get_authed_client
from ui.components import (
    section_header, kpi_row, alert,
    solicitud_card, historial_timeline, status_badge_html,
)

# ── Catálogos ─────────────────────────────────────────────────────────────────
FASES = [
    "Nuevo", "Capacitación", "Planteamiento",
    "Desarrollo", "Pruebas", "Entrega", "Concluido", "Cancelado",
]
FASES_ACTIVAS    = {"Nuevo", "Capacitación", "Planteamiento", "Desarrollo", "Pruebas", "Entrega"}
FASES_EN_PROCESO = {"Capacitación", "Planteamiento", "Desarrollo", "Pruebas", "Entrega", "En Proceso"}
FASES_CERRADAS   = {"Concluido", "Cancelado"}

ANALYSTS = ["Sin asignar", "Abel", "Sasha", "Adrian", "Heidi"]

# Columnas que van a la tabla / Excel (en orden)
COLS_TABLA = [
    ("id",           "ID"),
    ("created_at",   "Fecha creación"),
    ("updated_at",   "Última actualización"),
    ("solicitante",  "Solicitante"),
    ("correo",       "Correo"),
    ("empresa",      "Empresa"),
    ("titulo",       "Título"),
    ("categoria",    "Categoría"),
    ("departamento", "Departamento"),
    ("prioridad",    "Prioridad"),
    ("estatus",      "Estatus"),
    ("assigned_to",  "Asignado a"),
    ("descripcion",  "Descripción"),
]


# ─────────────────────────────────────────────────────────────────────────────
# Helpers
# ─────────────────────────────────────────────────────────────────────────────
def _now() -> str:
    return datetime.now(timezone.utc).isoformat(timespec="seconds")


def _gestor_nombre() -> str:
    u      = current_user() or {}
    perfil = st.session_state.get("profile") or {}
    return perfil.get("full_name") or u.get("email", "Gestor")


def _get_tickets(limit: int = 1000) -> list:
    try:
        sb  = get_authed_client()
        res = (
            sb.table("tickets")
            .select("*")
            .order("created_at", desc=True)
            .limit(limit)
            .execute()
        )
        return res.data or []
    except Exception:
        return []


def _update_ticket(ticket_id: int, changes: dict) -> bool:
    try:
        get_authed_client().table("tickets").update(changes).eq("id", ticket_id).execute()
        return True
    except Exception as e:
        st.error(f"Error al guardar: {e}")
        return False


def _to_excel(df: pd.DataFrame) -> bytes:
    """Convierte un DataFrame a bytes de Excel con formato básico."""
    buf = BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Tickets")
        ws = writer.sheets["Tickets"]

        # Ancho automático por columna
        for col_cells in ws.columns:
            max_len = max(
                len(str(cell.value)) if cell.value is not None else 0
                for cell in col_cells
            )
            ws.column_dimensions[col_cells[0].column_letter].width = min(max_len + 4, 50)

        # Estilo encabezado
        from openpyxl.styles import Font, PatternFill, Alignment
        header_fill = PatternFill("solid", fgColor="1B2266")
        header_font = Font(color="FFFFFF", bold=True)
        for cell in ws[1]:
            cell.fill      = header_fill
            cell.font      = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

    return buf.getvalue()


# ─────────────────────────────────────────────────────────────────────────────
# Modal de edición
# ─────────────────────────────────────────────────────────────────────────────
@st.dialog("Actualizar ticket", width="large")
def _modal_edicion(ticket: dict, gestor: str):
    tid   = ticket.get("id")
    est   = ticket.get("estatus", "Nuevo")
    badge = status_badge_html(est)

    st.markdown(
        f"**#{tid}** — {ticket.get('titulo','')}&nbsp;&nbsp;{badge}",
        unsafe_allow_html=True,
    )
    st.caption(
        f"Solicitante: {ticket.get('solicitante','')}  |  "
        f"Creado: {str(ticket.get('created_at',''))[:10]}"
    )

    col1, col2, col3 = st.columns(3)
    col1.markdown(f"**Empresa:** {ticket.get('empresa','')}")
    col2.markdown(f"**Categoría:** {ticket.get('categoria','')}")
    col3.markdown(f"**Departamento:** {ticket.get('departamento','')}")

    desc = ticket.get("descripcion", "")
    if desc:
        with st.expander("Ver descripción original"):
            st.write(desc)

    st.divider()
    st.markdown("**📋 Historial:**")
    historial_timeline(ticket.get("historial") or [])
    st.divider()

    st.markdown("**✏️ Registrar actualización:**")
    col_a, col_b = st.columns(2)
    with col_a:
        idx_est    = FASES.index(est) if est in FASES else 0
        nueva_fase = st.selectbox("Nueva fase / estatus", FASES, index=idx_est, key=f"md_fase_{tid}")
    with col_b:
        asig_actual = ticket.get("assigned_to", "Sin asignar")
        idx_asig    = ANALYSTS.index(asig_actual) if asig_actual in ANALYSTS else 0
        nuevo_asig  = st.selectbox("Asignado a", ANALYSTS, index=idx_asig, key=f"md_asig_{tid}")

    comentario = st.text_area(
        "Comentario / nota",
        placeholder="Describe el avance, acuerdo o ajuste...",
        height=110, key=f"md_com_{tid}",
    )

    col_btn1, col_btn2 = st.columns([1, 3])
    with col_btn1:
        guardar = st.button("💾 Guardar", type="primary", key=f"md_save_{tid}")
    with col_btn2:
        if st.button("✖️ Cancelar", key=f"md_cancel_{tid}"):
            st.rerun()

    if guardar:
        cambios_texto = []
        if nueva_fase != est:
            cambios_texto.append(f"Fase: {est} → {nueva_fase}")
        if nuevo_asig != asig_actual:
            cambios_texto.append(f"Asignado: {asig_actual} → {nuevo_asig}")
        if comentario.strip():
            cambios_texto.append(f"Comentario: {comentario.strip()}")

        if not cambios_texto:
            st.warning("No hay cambios que guardar.")
            return

        now         = _now()
        historial   = list(ticket.get("historial") or [])
        action_type = "comment" if comentario.strip() and nueva_fase == est else "fase"
        historial.append({
            "at": now, "by": gestor,
            "action": action_type,
            "details": " | ".join(cambios_texto),
        })

        changes = {
            "updated_at": now, "updated_by": gestor,
            "historial": historial, "assigned_to": nuevo_asig,
        }
        if nueva_fase != est:
            changes["estatus"] = nueva_fase

        if _update_ticket(tid, changes):
            st.success("✅ Cambios guardados.")
            st.cache_data.clear()
            st.rerun()


# ─────────────────────────────────────────────────────────────────────────────
# Badge estatus coloreado para la tabla (HTML)
# ─────────────────────────────────────────────────────────────────────────────
_BADGE_COLORS = {
    "Nuevo":         ("#1D4ED8", "#EFF6FF"),
    "Capacitación":  ("#7C3AED", "#F5F3FF"),
    "Planteamiento": ("#0891B2", "#ECFEFF"),
    "Desarrollo":    ("#D97706", "#FFFBEB"),
    "Pruebas":       ("#EA580C", "#FFF7ED"),
    "Entrega":       ("#059669", "#ECFDF5"),
    "Concluido":     ("#16A34A", "#F0FDF4"),
    "Cancelado":     ("#DC2626", "#FEF2F2"),
    "En Proceso":    ("#D97706", "#FFFBEB"),
}
_BADGE_DEFAULT = ("#6B7280", "#F9FAFB")

def _badge_html(est: str) -> str:
    color, bg = _BADGE_COLORS.get(est, _BADGE_DEFAULT)
    return (
        f'<span style="background:{bg};color:{color};border:1px solid {color};'
        f'border-radius:12px;padding:2px 10px;font-size:0.75rem;font-weight:700;">'
        f'{est}</span>'
    )


# ─────────────────────────────────────────────────────────────────────────────
# RENDER PRINCIPAL
# ─────────────────────────────────────────────────────────────────────────────
def render():
    section_header("🎫", "Gestión de Tickets",
                   "Administra fases, asignaciones y comentarios")

    gestor  = _gestor_nombre()
    tickets = _get_tickets()

    if not tickets:
        alert("info", "No hay tickets registrados.")
        return

    # ── ① KPIs ───────────────────────────────────────────────────────────────
    conteo = Counter(t.get("estatus", "Nuevo") for t in tickets)
    kpi_row([
        dict(icono="🆕", label="Nuevos",
             valor=conteo.get("Nuevo", 0), color="#1D4ED8"),
        dict(icono="⏳", label="En proceso",
             valor=sum(conteo.get(f, 0) for f in FASES_EN_PROCESO), color="#D97706"),
        dict(icono="✅", label="Concluidos",
             valor=conteo.get("Concluido", 0), color="#059669"),
        dict(icono="🚫", label="Cancelados",
             valor=conteo.get("Cancelado", 0), color="#DC2626"),
    ])

    st.divider()

    # ── Abrir modal si hay ticket seleccionado ────────────────────────────────
    if "gc_modal_id" in st.session_state:
        tid_sel    = st.session_state.pop("gc_modal_id")
        ticket_sel = next((t for t in tickets if t.get("id") == tid_sel), None)
        if ticket_sel:
            _modal_edicion(ticket_sel, gestor)

    # ── ② TICKETS ACTIVOS ─────────────────────────────────────────────────────
    activos = [t for t in tickets if t.get("estatus","Nuevo") in FASES_ACTIVAS]

    if activos:
        st.markdown(f"#### 🔄 Tickets activos ({len(activos)})")

        for t in activos:
            tid   = t.get("id")
            est   = t.get("estatus", "Nuevo")
            fecha = str(t.get("created_at", ""))[:10]

            meta = [
                ("🏢", t.get("empresa", "")),
                ("📂", t.get("categoria", "")),
                ("👤 Sol.", t.get("solicitante", "")),
                ("🔧", t.get("assigned_to", "Sin asignar")),
            ]

            clicked = solicitud_card(
                id_label=f"#{tid}",
                titulo=t.get("titulo", "(Sin título)"),
                fecha=fecha,
                estatus=est,
                meta=meta,
                on_edit_key=f"gc_open_{tid}",
            )
            if clicked:
                st.session_state["gc_modal_id"] = tid
                st.rerun()
    else:
        st.info("No hay tickets activos en este momento. ✅")

    st.divider()

    # ── ③ HISTORIAL COMPLETO — filtros + tabla + descarga ────────────────────
    st.markdown("#### 📋 Historial completo")

    # Filtros en fila
    col1, col2, col3, col4 = st.columns(4)
    with col1:
        todas_fases   = ["Todos"] + FASES
        filtro_est    = st.selectbox("Fase / estatus", todas_fases, key="gc_h_est")
    with col2:
        todas_empresas = ["Todas"] + sorted(set(
            t.get("empresa","") for t in tickets if t.get("empresa")
        ))
        filtro_emp = st.selectbox("Empresa", todas_empresas, key="gc_h_emp")
    with col3:
        todas_fechas = st.date_input(
            "Desde", value=None, key="gc_h_desde",
            help="Filtrar desde esta fecha de creación",
        )
    with col4:
        filtro_q = st.text_input("Título o solicitante", placeholder="Buscar...", key="gc_h_q")

    # Aplicar filtros
    filtrados = list(tickets)
    if filtro_est != "Todos":
        filtrados = [t for t in filtrados if t.get("estatus") == filtro_est]
    if filtro_emp != "Todas":
        filtrados = [t for t in filtrados if t.get("empresa","") == filtro_emp]
    if todas_fechas:
        filtrados = [
            t for t in filtrados
            if str(t.get("created_at",""))[:10] >= str(todas_fechas)
        ]
    if filtro_q.strip():
        q = filtro_q.strip().lower()
        filtrados = [
            t for t in filtrados
            if q in str(t.get("titulo","")).lower()
            or q in str(t.get("solicitante","")).lower()
        ]

    st.caption(f"Mostrando {len(filtrados)} de {len(tickets)} ticket(s)")

    # Tabla formateada
    if filtrados:
        df = pd.DataFrame([
            {label: t.get(col, "") for col, label in COLS_TABLA}
            for t in filtrados
        ])

        # Formatear fechas
        for fecha_col in ["Fecha creación", "Última actualización"]:
            if fecha_col in df.columns:
                df[fecha_col] = df[fecha_col].astype(str).str[:10]

        # Renderizar tabla con estatus coloreado via HTML
        def _row_style(row):
            est = row.get("Estatus", "")
            color, bg = _BADGE_COLORS.get(est, _BADGE_DEFAULT)
            return [
                f"color: {color}; font-weight: 700;" if col == "Estatus" else ""
                for col in row.index
            ]

        st.dataframe(
            df,
            use_container_width=True,
            hide_index=True,
            column_config={
                "ID":        st.column_config.NumberColumn("ID", width="small"),
                "Estatus":   st.column_config.TextColumn("Estatus", width="medium"),
                "Título":    st.column_config.TextColumn("Título", width="large"),
                "Correo":    st.column_config.TextColumn("Correo", width="medium"),
                "Descripción": st.column_config.TextColumn("Descripción", width="large"),
            },
        )

        # Descarga Excel (solo lo filtrado)
        excel_bytes = _to_excel(df)
        st.download_button(
            label="⬇️ Descargar Excel",
            data=excel_bytes,
            file_name="tickets.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="gc_tick_dl_excel",
        )
    else:
        st.info("No hay resultados con los filtros aplicados.")
