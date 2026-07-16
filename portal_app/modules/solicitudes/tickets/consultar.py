# portal_app/modules/solicitudes/tickets/consultar.py
# ─────────────────────────────────────────────────────────────────────────────
# Vista "Mis Tickets" para el usuario final.
# Sin HTML propio — todo el visual viene de ui/components.
# ─────────────────────────────────────────────────────────────────────────────
from __future__ import annotations
from io import BytesIO

import pandas as pd
import streamlit as st

from .shared import log_accion
from services.supabase_client import current_user, get_authed_client
from ui.components import (
    section_header, kpi_row, alert,
    solicitud_card, historial_timeline, status_badge_html,
    solicitudes_table,
    ESTATUS_CFG,
)

FASES_EN_PROCESO = {"Capacitación", "Planteamiento", "Desarrollo", "Pruebas", "Entrega", "En Proceso"}

COLS_TABLA = [
    ("id",           "ID"),
    ("created_at",   "Fecha creación"),
    ("updated_at",   "Última actualización"),
    ("empresa",      "Empresa"),
    ("titulo",       "Título"),
    ("categoria",    "Categoría"),
    ("departamento", "Departamento"),
    ("prioridad",    "Prioridad"),
    ("estatus",      "Estatus"),
    ("assigned_to",  "Asignado a"),
    ("descripcion",  "Descripción"),
]


# ── Query ─────────────────────────────────────────────────────────────────────
@st.cache_data(ttl=30, show_spinner=False)
def _mis_tickets(user_email: str) -> list:
    try:
        sb = get_authed_client()
        res = (
            sb.table("tickets")
            .select("*")
            .ilike("correo", user_email)
            .order("created_at", desc=True)
            .limit(500)
            .execute()
        )
        return res.data or []
    except Exception:
        return []


# ── Excel ─────────────────────────────────────────────────────────────────────
def _to_excel(df: pd.DataFrame) -> bytes:
    buf = BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Mis Tickets")
        ws = writer.sheets["Mis Tickets"]
        for col_cells in ws.columns:
            max_len = max(
                len(str(c.value)) if c.value is not None else 0
                for c in col_cells
            )
            ws.column_dimensions[col_cells[0].column_letter].width = min(max_len + 4, 50)
        from openpyxl.styles import Font, PatternFill, Alignment
        fill = PatternFill("solid", fgColor="1B2266")
        font = Font(color="FFFFFF", bold=True)
        for cell in ws[1]:
            cell.fill      = fill
            cell.font      = font
            cell.alignment = Alignment(horizontal="center", vertical="center")
    return buf.getvalue()


# ── Modal de detalle (solo lectura) ───────────────────────────────────────────
@st.dialog("Detalle del ticket", width="large")
def _modal_detalle(ticket: dict):
    est   = ticket.get("estatus", "Nuevo")
    badge = status_badge_html(est)

    st.markdown(
        f"**#{ticket.get('id')}** — {ticket.get('titulo','')}&nbsp;&nbsp;{badge}",
        unsafe_allow_html=True,
    )
    st.caption(
        f"Creado: {str(ticket.get('created_at',''))[:10]}  |  "
        f"Actualizado: {str(ticket.get('updated_at',''))[:10]}"
    )

    col1, col2, col3 = st.columns(3)
    col1.markdown(f"**Empresa:** {ticket.get('empresa','')}")
    col2.markdown(f"**Categoría:** {ticket.get('categoria','')}")
    col3.markdown(f"**Departamento:** {ticket.get('departamento','')}")

    col4, col5 = st.columns(2)
    col4.markdown(f"**Prioridad:** {ticket.get('prioridad','Normal')}")
    col5.markdown(f"**Asignado a:** {ticket.get('assigned_to','Sin asignar')}")

    desc = ticket.get("descripcion", "")
    if desc:
        st.markdown("---")
        st.markdown(f"**Descripción:**  \n{desc}")

    st.markdown("---")
    st.markdown("**Historial de actividad:**")
    historial_timeline(ticket.get("historial") or [])


# ── Render principal ──────────────────────────────────────────────────────────
def render():
    u = current_user()
    if not u:
        alert("error", "Debes iniciar sesión para ver tus tickets.")
        st.stop()

    user_email = (u.get("email") or "").strip().lower()
    tickets    = _mis_tickets(user_email)

    section_header("🎫", "Mis tickets", "Consulta el estado de tus solicitudes")

    # ── KPIs ──────────────────────────────────────────────────────────────────
    counts: dict[str, int] = {}
    for t in tickets:
        s = t.get("estatus", "Nuevo")
        counts[s] = counts.get(s, 0) + 1

    kpi_row([
        dict(icono="🆕", label="Nuevos",
             valor=counts.get("Nuevo", 0), color="#1D4ED8"),
        dict(icono="⏳", label="En proceso",
             valor=sum(counts.get(f, 0) for f in FASES_EN_PROCESO), color="#D97706"),
        dict(icono="✅", label="Concluidos",
             valor=counts.get("Concluido", 0), color="#059669"),
        dict(icono="🚫", label="Cancelados",
             valor=counts.get("Cancelado", 0), color="#DC2626"),
    ])

    if not tickets:
        alert("info", "Aún no tienes tickets. Usa 'Crear Ticket' para enviar una solicitud.")
        return

    st.divider()

    # ── Filtros ───────────────────────────────────────────────────────────────
    col1, col2, col3 = st.columns(3)
    with col1:
        estatuses_presentes = sorted(set(t.get("estatus", "Nuevo") for t in tickets))
        filtro_est = st.selectbox("Estatus", ["Todos"] + estatuses_presentes, key="ctk_est")
    with col2:
        filtro_q = st.text_input(
            "Buscar título", placeholder="Ej. Reporte ventas", key="ctk_q"
        )
    with col3:
        orden = st.selectbox("Ordenar", ["Más recientes", "Más antiguos"], key="ctk_ord")

    filtrados = list(tickets)
    if filtro_est != "Todos":
        filtrados = [t for t in filtrados if t.get("estatus") == filtro_est]
    if filtro_q.strip():
        q = filtro_q.strip().lower()
        filtrados = [t for t in filtrados if q in str(t.get("titulo", "")).lower()]
    if orden == "Más antiguos":
        filtrados = list(reversed(filtrados))

    st.caption(f"Mostrando {len(filtrados)} ticket(s)")
    st.divider()

    # ── Modal ─────────────────────────────────────────────────────────────────
    if "ctk_modal_id" in st.session_state:
        tid = st.session_state.pop("ctk_modal_id")
        ticket_sel = next((t for t in tickets if t.get("id") == tid), None)
        if ticket_sel:
            _modal_detalle(ticket_sel)

    # ── Cards ─────────────────────────────────────────────────────────────────
    for t in filtrados:
        tid   = t.get("id")
        est   = t.get("estatus", "Nuevo")
        fecha = str(t.get("created_at", ""))[:10]

        meta = [
            ("🏢", t.get("empresa", "")),
            ("📂", t.get("categoria", "")),
            ("⚡", t.get("prioridad", "Normal")),
            ("👤", t.get("assigned_to", "Sin asignar")),
        ]

        clicked = solicitud_card(
            id_label=f"#{tid}",
            titulo=t.get("titulo", "(Sin título)"),
            fecha=fecha,
            estatus=est,
            meta=meta,
            on_edit_key=f"ctk_open_{tid}",
        )
        if clicked:
            st.session_state["ctk_modal_id"] = tid
            st.rerun()

    # ── Tabla + descarga Excel (al final, debajo de las cards) ─────────────────
    if filtrados:
        st.divider()
        st.markdown("#### 📊 Resumen descargable")

        df = pd.DataFrame([
            {label: t.get(col, "") for col, label in COLS_TABLA}
            for t in filtrados
        ])
        for fecha_col in ["Fecha creación", "Última actualización"]:
            if fecha_col in df.columns:
                df[fecha_col] = df[fecha_col].astype(str).str[:10]

        solicitudes_table(df)

        descargado_excel = st.download_button(
            "⬇️ Descargar Excel",
            data=_to_excel(df),
            file_name="mis_tickets.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="ctk_dl",
        )
        if descargado_excel:
            log_accion("exportar_excel", {"filas": len(df)})
