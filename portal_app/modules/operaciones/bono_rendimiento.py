# portal_app/modules/operaciones/bono_rendimiento.py
# ─────────────────────────────────────────────────────────────────────────────
# Calculadora de Bono por Rendimiento de Diesel
# Empresas: Picus (prefijo P) e Igloo (prefijo G)
#
# Lógica:
#   rendimiento_real  = km / litros_cargados
#   litros_a_gastar   = km / rendimiento_minimo  (base del catálogo)
#   litros_ahorrados  = litros_a_gastar - litros_reales  (si > 0 → bono)
#   litros_de_mas     = litros_reales - litros_a_gastar  (si > 0 → descuento)
#   $ a pagar         = litros_ahorrados * precio_diesel
#   $ a descontar     = litros_de_mas * precio_diesel
#
# Sin HTML propio — todo visual va a través de ui/components.
# ─────────────────────────────────────────────────────────────────────────────
from __future__ import annotations
from io import BytesIO

import pandas as pd
import streamlit as st

from ui.components import section_header, kpi_row, alert, divider

# ── Catálogo de unidades ──────────────────────────────────────────────────────
# Fuente: Hoja2 de los archivos Excel entregados por operaciones.
# Para actualizar: modificar directamente estas listas.
# Columnas: (unidad, placa, marca, modelo, año, motor, rend_esperado, rend_minimo)

_CATALOGO_PICUS: list[dict] = [
    {"unidad": "P00100", "placa": "594AP5",    "marca": "KENWORTH",        "modelo": "T660",        "año": 2012, "motor": "Cummins ISM",        "rend_esperado": 2.85, "rend_minimo": 2.75},
    {"unidad": "P00101", "placa": "195EH9",    "marca": "KENWORTH",        "modelo": "T660",        "año": 2012, "motor": "Cummins ISM",        "rend_esperado": 2.85, "rend_minimo": 2.75},
    {"unidad": "P00102", "placa": "94AD7Z",    "marca": "KENWORTH",        "modelo": "T660",        "año": 2012, "motor": "Cummins ISM",        "rend_esperado": 2.85, "rend_minimo": 2.75},
    {"unidad": "P00103", "placa": "591AP5",    "marca": "KENWORTH",        "modelo": "T660",        "año": 2012, "motor": "Cummins ISM",        "rend_esperado": 2.85, "rend_minimo": 2.75},
    {"unidad": "P00104", "placa": "45AG1L",    "marca": "KENWORTH",        "modelo": "T660",        "año": 2012, "motor": "Cummins ISM",        "rend_esperado": 2.85, "rend_minimo": 2.75},
    {"unidad": "P00105", "placa": "599AP5",    "marca": "KENWORTH",        "modelo": "T660",        "año": 2012, "motor": "Cummins ISM",        "rend_esperado": 2.85, "rend_minimo": 2.75},
    {"unidad": "P00106", "placa": "598AP5",    "marca": "KENWORTH",        "modelo": "T660",        "año": 2012, "motor": "Cummins ISM",        "rend_esperado": 2.85, "rend_minimo": 2.75},
    {"unidad": "P00107", "placa": "76AK5V",    "marca": "KENWORTH",        "modelo": "T660",        "año": 2012, "motor": "Cummins ISM",        "rend_esperado": 2.85, "rend_minimo": 2.75},
    {"unidad": "P00108", "placa": "81AK6V",    "marca": "KENWORTH",        "modelo": "T660",        "año": 2012, "motor": "Cummins ISM",        "rend_esperado": 2.85, "rend_minimo": 2.75},
    {"unidad": "P00109", "placa": "12AL5Y",    "marca": "KENWORTH",        "modelo": "T660",        "año": 2012, "motor": "Cummins ISM",        "rend_esperado": 2.85, "rend_minimo": 2.75},
    {"unidad": "P00110", "placa": "551EV2",    "marca": "KENWORTH",        "modelo": "T660",        "año": 2012, "motor": "Cummins ISM",        "rend_esperado": 2.85, "rend_minimo": 2.75},
    {"unidad": "P00111", "placa": "44AG1L",    "marca": "KENWORTH",        "modelo": "T660",        "año": 2012, "motor": "Cummins ISM",        "rend_esperado": 2.85, "rend_minimo": 2.75},
    {"unidad": "P00113", "placa": "580AP5",    "marca": "KENWORTH",        "modelo": "T660",        "año": 2012, "motor": "Cummins ISM",        "rend_esperado": 2.85, "rend_minimo": 2.75},
    {"unidad": "P00114", "placa": "579AP5",    "marca": "KENWORTH",        "modelo": "T660",        "año": 2012, "motor": "Cummins ISM",        "rend_esperado": 2.85, "rend_minimo": 2.75},
    {"unidad": "P00115", "placa": "781EX6",    "marca": "KENWORTH",        "modelo": "T660",        "año": 2015, "motor": "Cummins ISM",        "rend_esperado": 2.85, "rend_minimo": 2.75},
    {"unidad": "P00116", "placa": "780EX6",    "marca": "KENWORTH",        "modelo": "T660",        "año": 2015, "motor": "Cummins ISM",        "rend_esperado": 2.85, "rend_minimo": 2.75},
    {"unidad": "P00117", "placa": "779EX6",    "marca": "KENWORTH",        "modelo": "T660",        "año": 2015, "motor": "Cummins ISM",        "rend_esperado": 2.85, "rend_minimo": 2.75},
    {"unidad": "P00118", "placa": "786EX6",    "marca": "KENWORTH",        "modelo": "T660",        "año": 2015, "motor": "Cummins ISM",        "rend_esperado": 2.85, "rend_minimo": 2.75},
    {"unidad": "P00119", "placa": "778EX6",    "marca": "KENWORTH",        "modelo": "T660",        "año": 2015, "motor": "Cummins ISM",        "rend_esperado": 2.85, "rend_minimo": 2.75},
    {"unidad": "P00120", "placa": "777EX6",    "marca": "KENWORTH",        "modelo": "T660",        "año": 2015, "motor": "Cummins ISM",        "rend_esperado": 2.85, "rend_minimo": 2.75},
    {"unidad": "P00122", "placa": "784EX6",    "marca": "KENWORTH",        "modelo": "T660",        "año": 2015, "motor": "Cummins ISM",        "rend_esperado": 2.85, "rend_minimo": 2.75},
    {"unidad": "P00123", "placa": "40AL3Y",    "marca": "KENWORTH",        "modelo": "T660",        "año": 2015, "motor": "Cummins ISM",        "rend_esperado": 2.85, "rend_minimo": 2.75},
    {"unidad": "P00125", "placa": "771EX6",    "marca": "KENWORTH",        "modelo": "T660",        "año": 2015, "motor": "Cummins ISM",        "rend_esperado": 2.85, "rend_minimo": 2.75},
    {"unidad": "P00126", "placa": "770EX6",    "marca": "KENWORTH",        "modelo": "T660",        "año": 2015, "motor": "Cummins ISM",        "rend_esperado": 2.85, "rend_minimo": 2.75},
    {"unidad": "P00127", "placa": "31AL2Y",    "marca": "KENWORTH",        "modelo": "T660",        "año": 2015, "motor": "Cummins ISM",        "rend_esperado": 2.85, "rend_minimo": 2.75},
    {"unidad": "P00128", "placa": "768EX6",    "marca": "KENWORTH",        "modelo": "T660",        "año": 2015, "motor": "Cummins ISM",        "rend_esperado": 2.85, "rend_minimo": 2.75},
    {"unidad": "P00129", "placa": "767EX6",    "marca": "KENWORTH",        "modelo": "T660",        "año": 2015, "motor": "Cummins ISM",        "rend_esperado": 2.85, "rend_minimo": 2.75},
    {"unidad": "P00131", "placa": "92AF1N",    "marca": "KENWORTH",        "modelo": "T660",        "año": 2015, "motor": "Cummins ISM",        "rend_esperado": 2.85, "rend_minimo": 2.75},
    {"unidad": "P00132", "placa": "774EX6",    "marca": "KENWORTH",        "modelo": "T660",        "año": 2015, "motor": "Cummins ISM",        "rend_esperado": 2.85, "rend_minimo": 2.75},
    {"unidad": "P00133", "placa": "773EX6",    "marca": "KENWORTH",        "modelo": "T660",        "año": 2015, "motor": "Cummins ISM",        "rend_esperado": 2.85, "rend_minimo": 2.75},
    {"unidad": "P00134", "placa": "772EX6",    "marca": "KENWORTH",        "modelo": "T660",        "año": 2015, "motor": "Cummins ISM",        "rend_esperado": 2.85, "rend_minimo": 2.75},
    {"unidad": "P00135", "placa": "35AG2B",    "marca": "KENWORTH",        "modelo": "T660",        "año": 2018, "motor": "Cummins ISX",        "rend_esperado": 2.85, "rend_minimo": 2.75},
    {"unidad": "P00136", "placa": "50AG1L",    "marca": "KENWORTH",        "modelo": "T660",        "año": 2018, "motor": "Cummins ISX",        "rend_esperado": 2.85, "rend_minimo": 2.75},
    {"unidad": "P00137", "placa": "52AG1L",    "marca": "KENWORTH",        "modelo": "T660",        "año": 2018, "motor": "Cummins ISX",        "rend_esperado": 2.85, "rend_minimo": 2.75},
    {"unidad": "P00138", "placa": "49AG1L",    "marca": "KENWORTH",        "modelo": "T660",        "año": 2018, "motor": "Cummins ISX",        "rend_esperado": 2.85, "rend_minimo": 2.75},
    {"unidad": "P00139", "placa": "78AG2B",    "marca": "KENWORTH",        "modelo": "T660",        "año": 2018, "motor": "Cummins ISX",        "rend_esperado": 2.85, "rend_minimo": 2.75},
    {"unidad": "P00140", "placa": "53AG1L",    "marca": "KENWORTH",        "modelo": "T660",        "año": 2018, "motor": "Cummins ISX",        "rend_esperado": 2.85, "rend_minimo": 2.75},
    {"unidad": "P00141", "placa": "55AG1L",    "marca": "KENWORTH",        "modelo": "T660",        "año": 2018, "motor": "Cummins ISX",        "rend_esperado": 2.85, "rend_minimo": 2.75},
    {"unidad": "P00143", "placa": "79AG2B",    "marca": "KENWORTH",        "modelo": "T660",        "año": 2018, "motor": "Cummins ISX",        "rend_esperado": 2.85, "rend_minimo": 2.75},
    {"unidad": "P00144", "placa": "80AG2B",    "marca": "KENWORTH",        "modelo": "T660",        "año": 2018, "motor": "Cummins ISX",        "rend_esperado": 2.85, "rend_minimo": 2.75},
    {"unidad": "P00145", "placa": "82AG2B",    "marca": "KENWORTH",        "modelo": "T660",        "año": 2018, "motor": "Cummins ISX",        "rend_esperado": 2.85, "rend_minimo": 2.75},
    {"unidad": "P00146", "placa": "73AG2B",    "marca": "KENWORTH",        "modelo": "T660",        "año": 2018, "motor": "Cummins ISX",        "rend_esperado": 2.85, "rend_minimo": 2.75},
    {"unidad": "P00147", "placa": "81AG2B",    "marca": "KENWORTH",        "modelo": "T660",        "año": 2018, "motor": "Cummins ISX",        "rend_esperado": 2.85, "rend_minimo": 2.75},
    {"unidad": "P00148", "placa": "56AG1L",    "marca": "KENWORTH",        "modelo": "T660",        "año": 2018, "motor": "Cummins ISX",        "rend_esperado": 2.85, "rend_minimo": 2.75},
    {"unidad": "P00150", "placa": "54AG1L",    "marca": "KENWORTH",        "modelo": "T660",        "año": 2018, "motor": "Cummins ISX",        "rend_esperado": 2.85, "rend_minimo": 2.75},
    {"unidad": "P00151", "placa": "74AG2B",    "marca": "KENWORTH",        "modelo": "T660",        "año": 2018, "motor": "Cummins ISX",        "rend_esperado": 2.85, "rend_minimo": 2.75},
    {"unidad": "P00152", "placa": "77AG2B",    "marca": "KENWORTH",        "modelo": "T660",        "año": 2018, "motor": "Cummins ISX",        "rend_esperado": 2.85, "rend_minimo": 2.75},
    {"unidad": "P00153", "placa": "75AG2B",    "marca": "KENWORTH",        "modelo": "T660",        "año": 2018, "motor": "Cummins ISX",        "rend_esperado": 2.85, "rend_minimo": 2.75},
    {"unidad": "P00154", "placa": "76AG2B",    "marca": "KENWORTH",        "modelo": "T660",        "año": 2018, "motor": "Cummins ISX",        "rend_esperado": 2.85, "rend_minimo": 2.75},
    {"unidad": "P00156", "placa": "55AM1L",    "marca": "SINOTRUK SITRAK", "modelo": "—",           "año": 2021, "motor": "MT 13.43-50 EURO V", "rend_esperado": 2.85, "rend_minimo": 2.75},
    {"unidad": "P00157", "placa": "Pendiente", "marca": "SINOTRUK SITRAK", "modelo": "—",           "año": 2021, "motor": "MT 13.43-50 EURO V", "rend_esperado": 2.85, "rend_minimo": 2.75},
    {"unidad": "P00158", "placa": "54AM1L",    "marca": "SINOTRUK SITRAK", "modelo": "—",           "año": 2021, "motor": "MT 13.43-50 EURO V", "rend_esperado": 2.85, "rend_minimo": 2.75},
    {"unidad": "P00159", "placa": "53AM1L",    "marca": "SINOTRUK SITRAK", "modelo": "—",           "año": 2021, "motor": "MT 13.43-50 EURO V", "rend_esperado": 2.85, "rend_minimo": 2.75},
    {"unidad": "P00160", "placa": "52AM1L",    "marca": "SINOTRUK SITRAK", "modelo": "—",           "año": 2021, "motor": "MT 13.43-50 EURO V", "rend_esperado": 2.85, "rend_minimo": 2.75},
    {"unidad": "P00161", "placa": "51AM1L",    "marca": "SINOTRUK SITRAK", "modelo": "—",           "año": 2021, "motor": "MT 13.43-50 EURO V", "rend_esperado": 2.85, "rend_minimo": 2.75},
    {"unidad": "P00162", "placa": "49AM1L",    "marca": "SINOTRUK SITRAK", "modelo": "—",           "año": 2021, "motor": "MT 13.43-50 EURO V", "rend_esperado": 2.85, "rend_minimo": 2.75},
    {"unidad": "P00163", "placa": "50AM1L",    "marca": "SINOTRUK SITRAK", "modelo": "—",           "año": 2021, "motor": "MT 13.43-50 EURO V", "rend_esperado": 2.85, "rend_minimo": 2.75},
    {"unidad": "P00164", "placa": "48AM1L",    "marca": "SINOTRUK SITRAK", "modelo": "—",           "año": 2021, "motor": "MT 13.43-50 EURO V", "rend_esperado": 2.85, "rend_minimo": 2.75},
    {"unidad": "P00165", "placa": "47AM1L",    "marca": "SINOTRUK SITRAK", "modelo": "—",           "año": 2021, "motor": "MT 13.43-50 EURO V", "rend_esperado": 2.85, "rend_minimo": 2.75},
    {"unidad": "P00166", "placa": "13AM2L",    "marca": "FREIGHTLINER",    "modelo": "Cascadia",    "año": 2021, "motor": "Detroit DD13",        "rend_esperado": 2.85, "rend_minimo": 2.75},
    {"unidad": "P00167", "placa": "12AM2L",    "marca": "FREIGHTLINER",    "modelo": "Cascadia",    "año": 2021, "motor": "Detroit DD13",        "rend_esperado": 3.00, "rend_minimo": 2.85},
    {"unidad": "P00168", "placa": "11AM2L",    "marca": "FREIGHTLINER",    "modelo": "Cascadia",    "año": 2021, "motor": "Detroit DD13",        "rend_esperado": 3.00, "rend_minimo": 2.85},
    {"unidad": "P00169", "placa": "10AM2L",    "marca": "FREIGHTLINER",    "modelo": "Cascadia",    "año": 2021, "motor": "Detroit DD13",        "rend_esperado": 3.00, "rend_minimo": 2.85},
    {"unidad": "P00170", "placa": "09AM2L",    "marca": "FREIGHTLINER",    "modelo": "Cascadia",    "año": 2021, "motor": "Detroit DD13",        "rend_esperado": 3.00, "rend_minimo": 2.85},
    {"unidad": "P00171", "placa": "08AM2L",    "marca": "FREIGHTLINER",    "modelo": "Cascadia",    "año": 2021, "motor": "Detroit DD13",        "rend_esperado": 3.00, "rend_minimo": 2.85},
    {"unidad": "P00172", "placa": "07AM2L",    "marca": "FREIGHTLINER",    "modelo": "Cascadia",    "año": 2021, "motor": "Detroit DD13",        "rend_esperado": 3.00, "rend_minimo": 2.85},
    {"unidad": "P00173", "placa": "06AM2L",    "marca": "FREIGHTLINER",    "modelo": "Cascadia",    "año": 2021, "motor": "Detroit DD13",        "rend_esperado": 3.00, "rend_minimo": 2.85},
    {"unidad": "P00174", "placa": "05AM2L",    "marca": "FREIGHTLINER",    "modelo": "Cascadia",    "año": 2021, "motor": "Detroit DD13",        "rend_esperado": 3.00, "rend_minimo": 2.85},
    {"unidad": "P00175", "placa": "04AM2L",    "marca": "FREIGHTLINER",    "modelo": "Cascadia",    "año": 2021, "motor": "Detroit DD13",        "rend_esperado": 3.00, "rend_minimo": 2.85},
    {"unidad": "P00176", "placa": "03AM2L",    "marca": "FREIGHTLINER",    "modelo": "Cascadia",    "año": 2021, "motor": "Detroit DD13",        "rend_esperado": 3.00, "rend_minimo": 2.85},
    {"unidad": "P00177", "placa": "02AM2L",    "marca": "FREIGHTLINER",    "modelo": "Cascadia",    "año": 2021, "motor": "Detroit DD13",        "rend_esperado": 3.00, "rend_minimo": 2.85},
    {"unidad": "P00178", "placa": "01AM2L",    "marca": "FREIGHTLINER",    "modelo": "Cascadia",    "año": 2021, "motor": "Detroit DD13",        "rend_esperado": 3.00, "rend_minimo": 2.85},
    {"unidad": "P00179", "placa": "99AM1L",    "marca": "FREIGHTLINER",    "modelo": "Cascadia",    "año": 2021, "motor": "Detroit DD13",        "rend_esperado": 3.00, "rend_minimo": 2.85},
    {"unidad": "P00180", "placa": "98AM1L",    "marca": "FREIGHTLINER",    "modelo": "Cascadia",    "año": 2021, "motor": "Detroit DD13",        "rend_esperado": 3.00, "rend_minimo": 2.85},
    {"unidad": "P00061", "placa": "78AL7K",    "marca": "KENWORTH",        "modelo": "T660",        "año": 2007, "motor": "Cummins ISM",        "rend_esperado": 3.00, "rend_minimo": 2.85},
    {"unidad": "P00062", "placa": "462AP5",    "marca": "KENWORTH",        "modelo": "T660",        "año": 2007, "motor": "Cummins ISM",        "rend_esperado": 3.00, "rend_minimo": 2.85},
    {"unidad": "P00063", "placa": "302DS2",    "marca": "KENWORTH",        "modelo": "T660",        "año": 2007, "motor": "Cummins ISM",        "rend_esperado": 3.00, "rend_minimo": 2.85},
    {"unidad": "P00064", "placa": "303DS2",    "marca": "KENWORTH",        "modelo": "T660",        "año": 2007, "motor": "Cummins ISM",        "rend_esperado": 3.00, "rend_minimo": 2.85},
    {"unidad": "P00066", "placa": "305DS2",    "marca": "KENWORTH",        "modelo": "T660",        "año": 2007, "motor": "Cummins ISM",        "rend_esperado": 3.00, "rend_minimo": 2.85},
    {"unidad": "P00067", "placa": "306DS2",    "marca": "KENWORTH",        "modelo": "T660",        "año": 2007, "motor": "Cummins ISM",        "rend_esperado": 3.00, "rend_minimo": 2.85},
    {"unidad": "P00070", "placa": "309DS2",    "marca": "KENWORTH",        "modelo": "T660",        "año": 2007, "motor": "Cummins ISM",        "rend_esperado": 3.00, "rend_minimo": 2.85},
    {"unidad": "P00072", "placa": "311DS2",    "marca": "KENWORTH",        "modelo": "T660",        "año": 2007, "motor": "Cummins ISM",        "rend_esperado": 3.00, "rend_minimo": 2.85},
    {"unidad": "P00073", "placa": "463AP5",    "marca": "KENWORTH",        "modelo": "T660",        "año": 2007, "motor": "Cummins ISM",        "rend_esperado": 3.00, "rend_minimo": 2.85},
    {"unidad": "P00075", "placa": "38AL5Y",    "marca": "KENWORTH",        "modelo": "T660",        "año": 2008, "motor": "Cummins ISM",        "rend_esperado": 3.00, "rend_minimo": 2.85},
    {"unidad": "P00079", "placa": "691DS2",    "marca": "KENWORTH",        "modelo": "T660",        "año": 2008, "motor": "Cummins ISM",        "rend_esperado": 3.00, "rend_minimo": 2.85},
    {"unidad": "P00080", "placa": "46AG1L",    "marca": "KENWORTH",        "modelo": "T660",        "año": 2008, "motor": "Cummins ISM",        "rend_esperado": 3.00, "rend_minimo": 2.85},
    {"unidad": "P00082", "placa": "717AP5",    "marca": "KENWORTH",        "modelo": "T660",        "año": 2008, "motor": "Cummins ISM",        "rend_esperado": 3.00, "rend_minimo": 2.85},
    {"unidad": "P00083", "placa": "699DS2",    "marca": "KENWORTH",        "modelo": "T660",        "año": 2008, "motor": "Cummins ISM",        "rend_esperado": 3.00, "rend_minimo": 2.85},
    {"unidad": "P00084", "placa": "692DS2",    "marca": "KENWORTH",        "modelo": "T660",        "año": 2008, "motor": "Cummins ISM",        "rend_esperado": 3.00, "rend_minimo": 2.85},
    {"unidad": "P00085", "placa": "812DS2",    "marca": "KENWORTH",        "modelo": "T660",        "año": 2008, "motor": "Cummins ISM",        "rend_esperado": 3.00, "rend_minimo": 2.85},
    {"unidad": "P00087", "placa": "805DS2",    "marca": "KENWORTH",        "modelo": "T660",        "año": 2008, "motor": "Cummins ISM",        "rend_esperado": 3.00, "rend_minimo": 2.85},
    {"unidad": "P00088", "placa": "804DS2",    "marca": "KENWORTH",        "modelo": "T660",        "año": 2008, "motor": "Cummins ISM",        "rend_esperado": 3.00, "rend_minimo": 2.85},
    {"unidad": "P00090", "placa": "810DS2",    "marca": "KENWORTH",        "modelo": "T660",        "año": 2008, "motor": "Cummins ISM",        "rend_esperado": 3.00, "rend_minimo": 2.85},
    {"unidad": "P00091", "placa": "43AG1L",    "marca": "KENWORTH",        "modelo": "T660",        "año": 2008, "motor": "Cummins ISM",        "rend_esperado": 3.00, "rend_minimo": 2.85},
    {"unidad": "P00093", "placa": "808DS2",    "marca": "KENWORTH",        "modelo": "T660",        "año": 2008, "motor": "Cummins ISM",        "rend_esperado": 3.00, "rend_minimo": 2.85},
    {"unidad": "P00094", "placa": "09AG2B",    "marca": "KENWORTH",        "modelo": "T660",        "año": 2008, "motor": "Cummins ISM",        "rend_esperado": 3.00, "rend_minimo": 2.85},
    {"unidad": "P00095", "placa": "51AG2L",    "marca": "KENWORTH",        "modelo": "T660",        "año": 2008, "motor": "Cummins ISM",        "rend_esperado": 3.00, "rend_minimo": 2.85},
    {"unidad": "P00097", "placa": "882DS2",    "marca": "KENWORTH",        "modelo": "T660",        "año": 2008, "motor": "Cummins ISM",        "rend_esperado": 3.00, "rend_minimo": 2.85},
    {"unidad": "P00098", "placa": "94AH3T",    "marca": "KENWORTH",        "modelo": "T660",        "año": 2008, "motor": "Cummins ISM",        "rend_esperado": 3.00, "rend_minimo": 2.85},
    {"unidad": "P00099", "placa": "142AP5",    "marca": "KENWORTH",        "modelo": "T660",        "año": 2008, "motor": "Cummins ISM",        "rend_esperado": 3.00, "rend_minimo": 2.85},
]

_CATALOGO_IGLOO: list[dict] = [
    {"unidad": "G00014", "placa": "815DS2", "marca": "KENWORTH",       "modelo": "T660",        "año": 2008, "motor": "ISM",     "rend_esperado": 2.75, "rend_minimo": 2.65},
    {"unidad": "G00022", "placa": "800EX1", "marca": "KENWORTH",       "modelo": "T660",        "año": 2008, "motor": "ISM",     "rend_esperado": 2.75, "rend_minimo": 2.65},
    {"unidad": "G00024", "placa": "573AP5", "marca": "KENWORTH",       "modelo": "T660",        "año": 2012, "motor": "ISM",     "rend_esperado": 2.75, "rend_minimo": 2.65},
    {"unidad": "G00027", "placa": "98AM2L", "marca": "KENWORTH",       "modelo": "T660",        "año": 2012, "motor": "ISM",     "rend_esperado": 2.75, "rend_minimo": 2.65},
    {"unidad": "G00028", "placa": "577AP5", "marca": "KENWORTH",       "modelo": "T660",        "año": 2012, "motor": "ISM",     "rend_esperado": 2.75, "rend_minimo": 2.65},
    {"unidad": "G00029", "placa": "585AP5", "marca": "KENWORTH",       "modelo": "T660",        "año": 2012, "motor": "ISM",     "rend_esperado": 2.75, "rend_minimo": 2.65},
    {"unidad": "G00032", "placa": "31AK5V", "marca": "KENWORTH",       "modelo": "T660",        "año": 2012, "motor": "ISM",     "rend_esperado": 2.75, "rend_minimo": 2.65},
    {"unidad": "G00033", "placa": "588AP5", "marca": "KENWORTH",       "modelo": "T660",        "año": 2012, "motor": "ISM",     "rend_esperado": 2.75, "rend_minimo": 2.65},
    {"unidad": "G00034", "placa": "49AK9J", "marca": "KENWORTH",       "modelo": "T660",        "año": 2015, "motor": "ISM",     "rend_esperado": 2.75, "rend_minimo": 2.65},
    {"unidad": "G00035", "placa": "27AF1N", "marca": "FREIGHT LINER",  "modelo": "Cascadia",    "año": 2016, "motor": "DD15",    "rend_esperado": 2.75, "rend_minimo": 2.65},
    {"unidad": "G00037", "placa": "26AG2B", "marca": "KENWORTH",       "modelo": "T680",        "año": 2018, "motor": "ISX",     "rend_esperado": 2.75, "rend_minimo": 2.65},
    {"unidad": "G00038", "placa": "28AG2B", "marca": "KENWORTH",       "modelo": "T680",        "año": 2018, "motor": "ISX",     "rend_esperado": 2.75, "rend_minimo": 2.65},
    {"unidad": "G00039", "placa": "27AG2B", "marca": "KENWORTH",       "modelo": "T680",        "año": 2018, "motor": "ISX",     "rend_esperado": 2.75, "rend_minimo": 2.65},
    {"unidad": "G00040", "placa": "68AG2B", "marca": "KENWORTH",       "modelo": "T680",        "año": 2018, "motor": "ISX",     "rend_esperado": 2.75, "rend_minimo": 2.65},
    {"unidad": "G00041", "placa": "69AG2B", "marca": "KENWORTH",       "modelo": "T680",        "año": 2018, "motor": "ISX",     "rend_esperado": 2.75, "rend_minimo": 2.65},
    {"unidad": "G00042", "placa": "70AG2B", "marca": "KENWORTH",       "modelo": "T680",        "año": 2018, "motor": "ISX",     "rend_esperado": 2.75, "rend_minimo": 2.65},
    {"unidad": "G00043", "placa": "63AG2B", "marca": "KENWORTH",       "modelo": "T680",        "año": 2018, "motor": "ISX",     "rend_esperado": 2.75, "rend_minimo": 2.65},
    {"unidad": "G00044", "placa": "62AG2B", "marca": "KENWORTH",       "modelo": "T680",        "año": 2018, "motor": "ISX",     "rend_esperado": 2.75, "rend_minimo": 2.65},
    {"unidad": "G00045", "placa": "61AG2B", "marca": "KENWORTH",       "modelo": "T680",        "año": 2018, "motor": "ISX",     "rend_esperado": 2.75, "rend_minimo": 2.65},
    {"unidad": "G00046", "placa": "66AG2B", "marca": "KENWORTH",       "modelo": "T680",        "año": 2018, "motor": "ISX",     "rend_esperado": 2.75, "rend_minimo": 2.65},
    {"unidad": "G00047", "placa": "65AG2B", "marca": "KENWORTH",       "modelo": "T680",        "año": 2018, "motor": "ISX",     "rend_esperado": 2.75, "rend_minimo": 2.65},
    {"unidad": "G00048", "placa": "64AG2B", "marca": "KENWORTH",       "modelo": "T680",        "año": 2018, "motor": "ISX",     "rend_esperado": 2.75, "rend_minimo": 2.65},
    {"unidad": "G00049", "placa": "67AG2B", "marca": "KENWORTH",       "modelo": "T680",        "año": 2018, "motor": "ISX",     "rend_esperado": 2.75, "rend_minimo": 2.65},
    {"unidad": "G00050", "placa": "32AG1L", "marca": "KENWORTH",       "modelo": "T680",        "año": 2018, "motor": "ISX",     "rend_esperado": 2.75, "rend_minimo": 2.65},
    {"unidad": "G00051", "placa": "76AJ4K", "marca": "KENWORTH",       "modelo": "T680",        "año": 2018, "motor": "ISX",     "rend_esperado": 2.75, "rend_minimo": 2.65},
    {"unidad": "G00052", "placa": "33AG1L", "marca": "KENWORTH",       "modelo": "T680",        "año": 2018, "motor": "ISX",     "rend_esperado": 2.75, "rend_minimo": 2.65},
    {"unidad": "G00053", "placa": "34AG1L", "marca": "KENWORTH",       "modelo": "T680",        "año": 2018, "motor": "ISX",     "rend_esperado": 2.75, "rend_minimo": 2.65},
    {"unidad": "G00054", "placa": "38AG1L", "marca": "KENWORTH",       "modelo": "T680",        "año": 2018, "motor": "ISX",     "rend_esperado": 2.75, "rend_minimo": 2.65},
    {"unidad": "G00055", "placa": "39AG1L", "marca": "KENWORTH",       "modelo": "T680",        "año": 2018, "motor": "ISX",     "rend_esperado": 2.75, "rend_minimo": 2.65},
    {"unidad": "G00056", "placa": "37AG1L", "marca": "KENWORTH",       "modelo": "T680",        "año": 2018, "motor": "ISX",     "rend_esperado": 2.75, "rend_minimo": 2.65},
    {"unidad": "G00057", "placa": "36AG1L", "marca": "KENWORTH",       "modelo": "T680",        "año": 2018, "motor": "ISX",     "rend_esperado": 2.75, "rend_minimo": 2.65},
    {"unidad": "G00058", "placa": "35AG1L", "marca": "KENWORTH",       "modelo": "T680",        "año": 2018, "motor": "ISX",     "rend_esperado": 2.75, "rend_minimo": 2.65},
    {"unidad": "G00059", "placa": "40AG1L", "marca": "FREIGHT LINER",  "modelo": "New Cascadia","año": 2019, "motor": "ISX",     "rend_esperado": 2.75, "rend_minimo": 2.65},
    {"unidad": "G00060", "placa": "40AG1L", "marca": "FREIGHT LINER",  "modelo": "New Cascadia","año": 2024, "motor": "DD13",    "rend_esperado": 2.90, "rend_minimo": 2.75},
    {"unidad": "G00061", "placa": "40AG1L", "marca": "FREIGHT LINER",  "modelo": "New Cascadia","año": 2024, "motor": "DD13",    "rend_esperado": 2.90, "rend_minimo": 2.75},
    {"unidad": "G00062", "placa": "40AG1L", "marca": "FREIGHT LINER",  "modelo": "New Cascadia","año": 2024, "motor": "DD13",    "rend_esperado": 2.90, "rend_minimo": 2.75},
    {"unidad": "G00063", "placa": "40AG1L", "marca": "FREIGHT LINER",  "modelo": "New Cascadia","año": 2024, "motor": "DD13",    "rend_esperado": 2.90, "rend_minimo": 2.75},
    {"unidad": "G00064", "placa": "40AG1L", "marca": "FREIGHT LINER",  "modelo": "New Cascadia","año": 2024, "motor": "DD13",    "rend_esperado": 2.90, "rend_minimo": 2.75},
    {"unidad": "G00065", "placa": "—",      "marca": "FREIGHT LINER",  "modelo": "New Cascadia","año": 2024, "motor": "DD13",    "rend_esperado": 2.90, "rend_minimo": 2.75},
    {"unidad": "G00066", "placa": "—",      "marca": "FREIGHT LINER",  "modelo": "New Cascadia","año": 2024, "motor": "DD13",    "rend_esperado": 2.90, "rend_minimo": 2.75},
    {"unidad": "G00067", "placa": "—",      "marca": "FREIGHT LINER",  "modelo": "New Cascadia","año": 2024, "motor": "DD13",    "rend_esperado": 2.90, "rend_minimo": 2.75},
    {"unidad": "G00068", "placa": "—",      "marca": "FREIGHT LINER",  "modelo": "New Cascadia","año": 2024, "motor": "DD13",    "rend_esperado": 2.90, "rend_minimo": 2.75},
    {"unidad": "G00069", "placa": "—",      "marca": "FREIGHT LINER",  "modelo": "New Cascadia","año": 2024, "motor": "DD13",    "rend_esperado": 2.90, "rend_minimo": 2.75},
    {"unidad": "G00070", "placa": "—",      "marca": "FREIGHT LINER",  "modelo": "New Cascadia","año": 2026, "motor": "DD13",    "rend_esperado": 2.90, "rend_minimo": 2.75},
    {"unidad": "G00071", "placa": "—",      "marca": "FREIGHT LINER",  "modelo": "New Cascadia","año": 2026, "motor": "DD13",    "rend_esperado": 2.90, "rend_minimo": 2.75},
    {"unidad": "G00072", "placa": "—",      "marca": "FREIGHT LINER",  "modelo": "New Cascadia","año": 2026, "motor": "DD13",    "rend_esperado": 2.90, "rend_minimo": 2.75},
    {"unidad": "G00073", "placa": "—",      "marca": "FREIGHT LINER",  "modelo": "New Cascadia","año": 2026, "motor": "DD13",    "rend_esperado": 2.90, "rend_minimo": 2.75},
    {"unidad": "G00074", "placa": "—",      "marca": "FREIGHT LINER",  "modelo": "New Cascadia","año": 2026, "motor": "DD13",    "rend_esperado": 2.90, "rend_minimo": 2.75},
    {"unidad": "G00075", "placa": "—",      "marca": "FREIGHT LINER",  "modelo": "New Cascadia","año": 2026, "motor": "DD13",    "rend_esperado": 2.90, "rend_minimo": 2.75},
    {"unidad": "G00076", "placa": "—",      "marca": "FREIGHT LINER",  "modelo": "New Cascadia","año": 2026, "motor": "DD13",    "rend_esperado": 2.90, "rend_minimo": 2.75},
    {"unidad": "G00077", "placa": "—",      "marca": "FREIGHT LINER",  "modelo": "New Cascadia","año": 2026, "motor": "DD13",    "rend_esperado": 2.90, "rend_minimo": 2.75},
    {"unidad": "G00078", "placa": "—",      "marca": "FREIGHT LINER",  "modelo": "New Cascadia","año": 2026, "motor": "DD13",    "rend_esperado": 2.90, "rend_minimo": 2.75},
    {"unidad": "G00079", "placa": "—",      "marca": "FREIGHT LINER",  "modelo": "New Cascadia","año": 2026, "motor": "DD13",    "rend_esperado": 2.90, "rend_minimo": 2.75},
    {"unidad": "G00080", "placa": "—",      "marca": "FREIGHT LINER",  "modelo": "New Cascadia","año": 2026, "motor": "DD13",    "rend_esperado": 2.90, "rend_minimo": 2.75},
]

# ── Índice por unidad (lookup rápido) ────────────────────────────────────────
_IDX_PICUS = {u["unidad"]: u for u in _CATALOGO_PICUS}
_IDX_IGLOO = {u["unidad"]: u for u in _CATALOGO_IGLOO}

CATALOGOS = {
    "Picus":  (_CATALOGO_PICUS, _IDX_PICUS),
    "Igloo":  (_CATALOGO_IGLOO, _IDX_IGLOO),
}


# ── Cálculo central ───────────────────────────────────────────────────────────
def calcular_bono(km: float, litros: float, rend_minimo: float, precio_diesel: float) -> dict:
    """
    Retorna todos los valores calculados del bono.
    Puro Python — sin Streamlit — para facilitar pruebas.
    """
    if litros <= 0 or km <= 0:
        return {}

    rend_real      = km / litros
    litros_gastar  = km / rend_minimo
    diferencia     = litros_gastar - litros          # positivo = ahorro, negativo = exceso

    litros_ahorrados = max(diferencia, 0.0)
    litros_de_mas    = max(-diferencia, 0.0)

    a_pagar    = litros_ahorrados * precio_diesel
    a_descontar = litros_de_mas * precio_diesel
    procede_bono = rend_real >= rend_minimo

    return {
        "rend_real":       round(rend_real, 6),
        "litros_gastar":   round(litros_gastar, 6),
        "litros_ahorrados": round(litros_ahorrados, 6),
        "litros_de_mas":   round(litros_de_mas, 6),
        "a_pagar":         round(a_pagar, 2),
        "a_descontar":     round(a_descontar, 2),
        "procede_bono":    procede_bono,
    }


# ── Exportar Excel ────────────────────────────────────────────────────────────
def _exportar_excel(resultados: list[dict], empresa: str, precio_diesel: float) -> bytes:
    rows = []
    for r in resultados:
        rows.append({
            "Empresa":             empresa,
            "Unidad":              r["unidad"],
            "Placa":               r["placa"],
            "Marca":               r["marca"],
            "Modelo":              r["modelo"],
            "Año":                 r["año"],
            "Kilómetros":          r["km"],
            "Litros Cargados":     r["litros"],
            "Rend. Real (km/L)":   r["rend_real"],
            "Rend. Mínimo (km/L)": r["rend_minimo"],
            "Litros a Gastar":     r["litros_gastar"],
            "Litros Ahorrados":    r["litros_ahorrados"],
            "Litros de Más":       r["litros_de_mas"],
            "Precio Diesel (MXN)": precio_diesel,
            "$ a Pagar":           r["a_pagar"],
            "$ a Descontar":       r["a_descontar"],
            "Procede Bono":        "Sí" if r["procede_bono"] else "No",
        })

    df = pd.DataFrame(rows)
    buf = BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Bono Rendimiento")
        ws = writer.sheets["Bono Rendimiento"]
        # Ancho automático
        for col_cells in ws.columns:
            max_len = max(len(str(c.value)) if c.value is not None else 0 for c in col_cells)
            ws.column_dimensions[col_cells[0].column_letter].width = min(max_len + 4, 30)
        # Encabezado con color PGL
        from openpyxl.styles import Font, PatternFill, Alignment
        fill = PatternFill("solid", fgColor="1B2266")
        font = Font(color="FFFFFF", bold=True)
        for cell in ws[1]:
            cell.fill = fill
            cell.font = font
            cell.alignment = Alignment(horizontal="center")
    return buf.getvalue()


# ── UI principal ──────────────────────────────────────────────────────────────
def render():
    section_header("⛽", "Bono por Rendimiento de Diesel",
                   "Calcula el bono o descuento por litros según el rendimiento real vs. mínimo de cada unidad")

    # ── Configuración general ─────────────────────────────────────────────────
    col_emp, col_diesel, _ = st.columns([1, 1, 2])

    with col_emp:
        empresa = st.selectbox(
            "Empresa",
            options=["Picus", "Igloo"],
            key="br_empresa",
        )

    with col_diesel:
        precio_diesel = st.number_input(
            "Precio diesel (MXN/L)",
            min_value=1.0,
            max_value=100.0,
            value=24.01,
            step=0.01,
            format="%.2f",
            key="br_precio_diesel",
        )

    catalogo, idx = CATALOGOS[empresa]
    opciones_unidades = [u["unidad"] for u in catalogo]

    divider()

    # ── Selector de modo ──────────────────────────────────────────────────────
    modo = st.radio(
        "Modo de captura",
        options=["Una unidad", "Múltiples unidades"],
        horizontal=True,
        key="br_modo",
    )

    divider()

    # ════════════════════════════════════════════════════════════════════════════
    # MODO: UNA UNIDAD
    # ════════════════════════════════════════════════════════════════════════════
    if modo == "Una unidad":
        col_u, col_km, col_lit = st.columns([1.5, 1, 1])

        with col_u:
            unidad_sel = st.selectbox(
                "Unidad",
                options=opciones_unidades,
                key="br_unidad_sel",
            )

        datos_u = idx.get(unidad_sel, {})

        with col_km:
            km = st.number_input(
                "Kilómetros",
                min_value=1.0,
                max_value=99999.0,
                value=2700.0,
                step=1.0,
                format="%.0f",
                key="br_km",
            )

        with col_lit:
            litros = st.number_input(
                "Litros cargados",
                min_value=1.0,
                max_value=99999.0,
                value=950.0,
                step=1.0,
                format="%.0f",
                key="br_litros",
            )

        # Info de la unidad seleccionada
        if datos_u:
            st.caption(
                f"📋 {datos_u['marca']} {datos_u['modelo']} {datos_u['año']} · "
                f"Motor: {datos_u['motor']} · "
                f"Rend. esperado: **{datos_u['rend_esperado']}** km/L · "
                f"Rend. mínimo: **{datos_u['rend_minimo']}** km/L"
            )

        if st.button("Calcular bono", type="primary", key="br_calcular_uno"):
            rend_min = datos_u.get("rend_minimo", 2.75)
            res = calcular_bono(km, litros, rend_min, precio_diesel)

            if not res:
                alert("error", "Verifica que km y litros sean mayores a 0.")
                return

            # KPIs resultado
            color_rend = "#059669" if res["procede_bono"] else "#DC2626"
            kpi_row([
                dict(icono="📏", label="Rendimiento real",  valor=f"{res['rend_real']:.4f} km/L",  sub=f"Mínimo: {rend_min} km/L",       color=color_rend),
                dict(icono="🪣", label="Litros a gastar",   valor=f"{res['litros_gastar']:.2f} L",  sub="Base para el cálculo",            color="#1B2266"),
                dict(icono="💧", label="Litros ahorrados",  valor=f"{res['litros_ahorrados']:.2f} L", sub="Litros por debajo del mínimo",  color="#059669"),
                dict(icono="🔺", label="Litros de más",     valor=f"{res['litros_de_mas']:.2f} L",  sub="Litros por encima del mínimo",   color="#DC2626"),
            ])

            if res["procede_bono"]:
                alert("success", f"✅ Procede bono — **${res['a_pagar']:,.2f} MXN** a pagar al operador")
            else:
                alert("warn", f"⚠️ No procede bono — Se descontarán **${res['a_descontar']:,.2f} MXN** por exceso de consumo")

    # ════════════════════════════════════════════════════════════════════════════
    # MODO: MÚLTIPLES UNIDADES
    # ════════════════════════════════════════════════════════════════════════════
    else:
        st.caption("Captura los datos de todas las unidades de la quincena y calcula todos los bonos de una vez.")

        # ── Estado de la tabla editable ───────────────────────────────────────
        if "br_filas" not in st.session_state or st.session_state.get("br_empresa_prev") != empresa:
            st.session_state["br_filas"] = [
                {"unidad": opciones_unidades[0], "km": 2700.0, "litros": 950.0}
            ]
            st.session_state["br_empresa_prev"] = empresa

        filas = st.session_state["br_filas"]

        # ── Tabla de captura ──────────────────────────────────────────────────
        for i, fila in enumerate(filas):
            c1, c2, c3, c4 = st.columns([2, 1, 1, 0.3])
            with c1:
                fila["unidad"] = st.selectbox(
                    f"Unidad {i + 1}",
                    options=opciones_unidades,
                    index=opciones_unidades.index(fila["unidad"]) if fila["unidad"] in opciones_unidades else 0,
                    key=f"br_mul_unidad_{i}",
                    label_visibility="collapsed" if i > 0 else "visible",
                )
            with c2:
                fila["km"] = st.number_input(
                    "Km" if i == 0 else f"km_{i}",
                    min_value=1.0, max_value=99999.0,
                    value=float(fila["km"]), step=1.0, format="%.0f",
                    key=f"br_mul_km_{i}",
                    label_visibility="collapsed" if i > 0 else "visible",
                )
            with c3:
                fila["litros"] = st.number_input(
                    "Litros" if i == 0 else f"litros_{i}",
                    min_value=1.0, max_value=99999.0,
                    value=float(fila["litros"]), step=1.0, format="%.0f",
                    key=f"br_mul_litros_{i}",
                    label_visibility="collapsed" if i > 0 else "visible",
                )
            with c4:
                # Botón eliminar (no en la primera fila si solo hay una)
                if len(filas) > 1:
                    if st.button("✕", key=f"br_del_{i}", help="Eliminar fila"):
                        st.session_state["br_filas"].pop(i)
                        st.rerun()

        # ── Controles de filas ─────────────────────────────────────────────────
        col_add, col_calc, col_reset = st.columns([1, 1, 1])
        with col_add:
            if st.button("➕ Agregar unidad", key="br_add_fila", use_container_width=True):
                st.session_state["br_filas"].append(
                    {"unidad": opciones_unidades[0], "km": 2700.0, "litros": 950.0}
                )
                st.rerun()

        with col_reset:
            if st.button("🗑️ Limpiar todo", key="br_reset", use_container_width=True):
                st.session_state["br_filas"] = [
                    {"unidad": opciones_unidades[0], "km": 2700.0, "litros": 950.0}
                ]
                st.rerun()

        with col_calc:
            calcular = st.button("⚡ Calcular todos", type="primary",
                                 key="br_calcular_multi", use_container_width=True)

        # ── Resultados ────────────────────────────────────────────────────────
        if calcular:
            resultados = []
            for fila in filas:
                datos_u = idx.get(fila["unidad"], {})
                rend_min = datos_u.get("rend_minimo", 2.75)
                res = calcular_bono(fila["km"], fila["litros"], rend_min, precio_diesel)
                if res:
                    resultados.append({
                        "unidad":    fila["unidad"],
                        "placa":     datos_u.get("placa", "—"),
                        "marca":     datos_u.get("marca", "—"),
                        "modelo":    datos_u.get("modelo", "—"),
                        "año":       datos_u.get("año", "—"),
                        "km":        fila["km"],
                        "litros":    fila["litros"],
                        "rend_minimo": rend_min,
                        **res,
                    })

            if not resultados:
                alert("error", "No hay datos válidos para calcular.")
                return

            st.session_state["br_resultados"] = resultados
            st.session_state["br_empresa_res"] = empresa
            st.session_state["br_precio_res"] = precio_diesel

        # ── Mostrar tabla de resultados (persiste tras calcular) ──────────────
        if "br_resultados" in st.session_state:
            resultados = st.session_state["br_resultados"]
            divider()
            section_header("📊", "Resultados del cálculo")

            # KPIs resumen
            total_pagar    = sum(r["a_pagar"] for r in resultados)
            total_descontar = sum(r["a_descontar"] for r in resultados)
            con_bono       = sum(1 for r in resultados if r["procede_bono"])
            sin_bono       = len(resultados) - con_bono

            kpi_row([
                dict(icono="🚛", label="Unidades procesadas", valor=len(resultados),       sub="en este cálculo",              color="#1B2266"),
                dict(icono="✅", label="Con bono",            valor=con_bono,               sub="rendimiento sobre mínimo",     color="#059669"),
                dict(icono="⚠️", label="Sin bono / descuento", valor=sin_bono,             sub="rendimiento bajo mínimo",      color="#DC2626"),
                dict(icono="💰", label="Total a pagar",       valor=f"${total_pagar:,.2f}", sub="MXN acumulado",                color="#059669"),
                dict(icono="🔻", label="Total a descontar",   valor=f"${total_descontar:,.2f}", sub="MXN acumulado",             color="#DC2626"),
            ])

            # Tabla de resultados
            df_res = pd.DataFrame([{
                "Unidad":           r["unidad"],
                "Placa":            r["placa"],
                "Km":               r["km"],
                "Litros":           r["litros"],
                "Rend. Real":       round(r["rend_real"], 4),
                "Rend. Mínimo":     r["rend_minimo"],
                "Litros Ahorrados": round(r["litros_ahorrados"], 2),
                "Litros de Más":    round(r["litros_de_mas"], 2),
                "$ a Pagar":        r["a_pagar"],
                "$ a Descontar":    r["a_descontar"],
                "Bono":             "✅ Sí" if r["procede_bono"] else "❌ No",
            } for r in resultados])

            st.dataframe(
                df_res,
                use_container_width=True,
                hide_index=True,
                column_config={
                    "$ a Pagar":    st.column_config.NumberColumn(format="$%.2f"),
                    "$ a Descontar": st.column_config.NumberColumn(format="$%.2f"),
                },
            )

            # Descarga Excel
            excel_bytes = _exportar_excel(
                resultados,
                st.session_state.get("br_empresa_res", empresa),
                st.session_state.get("br_precio_res", precio_diesel),
            )
            st.download_button(
                label="⬇️ Descargar Excel",
                data=excel_bytes,
                file_name=f"bono_rendimiento_{empresa.lower()}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="br_download",
            )
