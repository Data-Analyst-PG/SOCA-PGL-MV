# portal_app/modules/auditoria/prorrateador_historico.py
import re
import pandas as pd
import streamlit as st
import openpyxl

from .shared import to_excel_bytes_sheets


MESES_ES = [
    "ENERO","FEBRERO","MARZO","ABRIL","MAYO","JUNIO",
    "JULIO","AGOSTO","SEPTIEMBRE","SETIEMBRE","OCTUBRE","NOVIEMBRE","DICIEMBRE"
]

HEADS = ["FACTURACIÓN","COSTOS DIRECTOS","UTILIDAD","% UT BRUTA","COSTOS INDIRECTOS","% CI",
         "GASTOS GENERALES","% GN","UT/PER","%UT/PER"]


def _es_titulo_mes(v) -> bool:
    if not isinstance(v, str):
        return False
    s = v.strip().upper()
    return any(m in s for m in MESES_ES) and re.search(r"\b20\d{2}\b", s)


def _extraer_mes(v) -> str:
    s = str(v).strip().upper()
    for m in MESES_ES:
        if m in s:
            return "SEPTIEMBRE" if m == "SETIEMBRE" else m
    return s


def _orden_mes(m):
    orden = ["ENERO","FEBRERO","MARZO","ABRIL","MAYO","JUNIO","JULIO","AGOSTO","SEPTIEMBRE","OCTUBRE","NOVIEMBRE","DICIEMBRE"]
    d = {mm: i+1 for i, mm in enumerate(orden)}
    return d.get(str(m).strip().upper(), 999)


def _normaliza_header(h):
    s = str(h).strip().upper()
    s = s.replace("FACTURACION", "FACTURACIÓN")
    return s


def parse_sheet(ws, sheet_name: str) -> pd.DataFrame:
    rows = []
    max_r = ws.max_row
    max_c = ws.max_column

    for r in range(1, max_r + 1):
        for c in range(1, max_c + 1):
            v = ws.cell(r, c).value
            if not _es_titulo_mes(v):
                continue

            mes = _extraer_mes(v)
            r_head = r + 1
            r_val = r + 2
            if r_val > max_r:
                continue

            headers = [ws.cell(r_head, c+i).value for i in range(0, 10)]
            values  = [ws.cell(r_val,  c+i).value for i in range(0, 10)]
            headers = [_normaliza_header(h) for h in headers]

            if "FACTURACIÓN" not in headers and "FACTURACION" not in headers:
                continue
            if "UT/PER" not in headers:
                continue

            d = {"Mes": mes, "Sucursal": str(sheet_name).strip().upper()}
            for h, val in zip(headers, values):
                d[h] = val

            out = {
                "Mes": d.get("Mes"),
                "Sucursal": d.get("Sucursal"),
                "Facturación": float(d.get("FACTURACIÓN") or 0),
                "Costos Dire": float(d.get("COSTOS DIRECTOS") or 0),
                "Utilidad": float(d.get("UTILIDAD") or 0),
                "% Ut Bruta": float(d.get("% UT BRUTA") or 0),
                "Costos Indirectos": float(d.get("COSTOS INDIRECTOS") or 0),
                "% CI": float(d.get("% CI") or 0),
                "Gastos Generales": float(d.get("GASTOS GENERALES") or 0),
                "% GN": float(d.get("% GN") or 0),
                "UT/PER": float(d.get("UT/PER") or 0),
                "%UT/PER": float(d.get("%UT/PER") or 0),
            }
            rows.append(out)

    if not rows:
        return pd.DataFrame()

    df = pd.DataFrame(rows).drop_duplicates(subset=["Mes","Sucursal"], keep="last")
    df["__m"] = df["Mes"].apply(_orden_mes)
    df = df.sort_values(["Sucursal","__m"]).drop(columns="__m").reset_index(drop=True)

    cols = ["Mes","Sucursal","Facturación","Costos Dire","Utilidad","% Ut Bruta",
            "Costos Indirectos","% CI","Gastos Generales","% GN","UT/PER","%UT/PER"]
    return df[cols]


def render():
    from ui.components import section_header, alert, divider
    section_header("📚", "Consolidar histórico (archivo con tablitas por mes)")
    st.caption("Sube el Excel histórico: una hoja por sucursal con bloques mensuales tipo 'ENERO 2025'.")

    archivo_hist = st.file_uploader(
        "Excel histórico",
        type=["xlsx"],
        key="historico_excel"
    )

    if not archivo_hist:
        alert("info", "Sube el archivo para detectar y consolidar bloques por mes.")
        return

    try:
        wb = openpyxl.load_workbook(archivo_hist, data_only=True)

        dfs = {}
        for sh in wb.sheetnames:
            ws = wb[sh]
            df_sh = parse_sheet(ws, sh)
            if not df_sh.empty:
                dfs[sh] = df_sh

        if not dfs:
            alert("error", "No encontré bloques tipo 'ENERO 2025'. Revisa el formato del archivo.")
            return

        st.success(f"Listo: detecté consolidado en {len(dfs)} hojas.")
        sucursales = sorted(dfs.keys())
        suc_sel = st.selectbox("Ver consolidado de sucursal", sucursales, key="ver_suc_hist")
        st.dataframe(dfs[suc_sel], use_container_width=True)

        st.download_button(
            "📥 Descargar consolidado (1 hoja por sucursal)",
            data=to_excel_bytes_sheets(dfs),
            file_name="consolidado_historico_por_sucursal.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    except Exception as e:
        st.error(f"Error procesando histórico: {e}")
        st.exception(e)
